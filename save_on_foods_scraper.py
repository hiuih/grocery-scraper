#!/usr/bin/env python3
"""
SAVE-ON-FOODS PRICE SCRAPER
================================
Scrapes every product from Save-On-Foods (store rsid 1982) and saves
results to Save_On_Foods_Products.xlsx on your Desktop.

Save-On-Foods sits behind Cloudflare bot protection, so this script uses
`patchright` (a stealth Playwright fork) and retries fresh browser
sessions until one gets past the challenge. This is not 100% guaranteed
on every run, so the script retries automatically at multiple levels.

HOW TO RUN:
    python3 save_on_foods_scraper.py
  Takes roughly 1-2 hours (headless retries make it slower than Fresh St).
  Excel file appears on your Desktop when done.
"""

import sys, os

try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def p(msg="", flush=True):
    print(msg, flush=flush)


import subprocess


def pip_install(pkg):
    r = subprocess.run([sys.executable, "-m", "pip", "install", pkg, "-q"],
                       capture_output=True, text=True)
    if r.returncode != 0:
        subprocess.run([sys.executable, "-m", "pip", "install", pkg])


p("=" * 60)
p("  SAVE-ON-FOODS PRICE SCRAPER")
p("=" * 60)
p()
p("Checking dependencies...")

for pkg in ["patchright", "openpyxl"]:
    try:
        __import__(pkg)
        p(f"  [OK] {pkg}")
    except ImportError:
        p(f"  Installing {pkg}...")
        pip_install(pkg)
        p(f"  [OK] {pkg} installed.")

p("  Checking browser...")
subprocess.run([sys.executable, "-m", "patchright", "install", "chromium"],
               capture_output=True, text=True)
p("  [OK] Browser ready.")
p()

try:
    from patchright.sync_api import sync_playwright
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import re, time
    from datetime import datetime
except ImportError as e:
    p(f"  ERROR: {e}")
    try:
        input("  Press ENTER to close...")
    except EOFError:
        pass
    sys.exit(1)


def find_desktop():
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
        d, _ = winreg.QueryValueEx(key, "Desktop")
        winreg.CloseKey(key)
        if os.path.isdir(d):
            return d
    except Exception:
        pass
    home = os.path.expanduser("~")
    for candidate in [
        os.path.join(home, "Desktop"),
        os.path.join(home, "OneDrive", "Desktop"),
        os.path.join(home, "OneDrive - Personal", "Desktop"),
    ]:
        if os.path.isdir(candidate):
            return candidate
    fb = os.path.join(home, "Desktop")
    os.makedirs(fb, exist_ok=True)
    return fb


OUTPUT_DIR    = find_desktop()
SAVEON_FILE   = os.path.join(OUTPUT_DIR, "Save_On_Foods_Products.xlsx")
STATE_FILE    = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".saveon_session.json")

SAVEON_BASE = "https://www.saveonfoods.com/sm/pickup/rsid/1982"

CARD_SEL  = "[data-testid^='ProductCardWrapper']"
PRICE_SEL = "[class*='ProductCardPrice--']"
WAS_SEL   = "[class*='WasPrice--']"

USER_AGENT = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
              "AppleWebKit/537.36 (KHTML, like Gecko) "
              "Chrome/124.0.0.0 Safari/537.36")

CHALLENGE_RETRIES = 3
CHALLENGE_TEXT_MARKERS = ("security verification", "just a moment", "checking your browser")


def wait_for_real_content(page, max_wait=25):
    """Poll body text: Cloudflare's challenge page renders almost-empty (~8 chars) or
    shows its own verification copy (~267 chars) before the real app hydrates."""
    start = time.time()
    txt = ""
    while time.time() - start < max_wait:
        try:
            txt = (page.evaluate("document.body.innerText") or "").strip()
        except Exception:
            txt = ""
        if len(txt) > 50 and not any(m in txt.lower() for m in CHALLENGE_TEXT_MARKERS):
            return True
        time.sleep(1)
    return False


def is_challenged(page):
    return not wait_for_real_content(page, max_wait=25)


def headed_clear_and_save(playwright, timeout=45000):
    """Open a REAL, visible browser window so Cloudflare's challenge resolves the
    way it does for an ordinary human visitor, then save the resulting session
    cookies to disk so the rest of the run can proceed headlessly."""
    p("  Opening a visible browser window to clear Cloudflare's check...")
    p("  (If a checkbox or puzzle appears, solve it — this should only happen once.)")
    browser = playwright.chromium.launch(headless=False)
    ctx = browser.new_context(
        user_agent=USER_AGENT,
        viewport={"width": 1280, "height": 900},
        locale="en-CA",
    )
    page = ctx.new_page()
    cleared = False
    try:
        page.goto(SAVEON_BASE, wait_until="domcontentloaded", timeout=timeout)
        for attempt in range(6):
            if wait_for_real_content(page, max_wait=20):
                cleared = True
                break
            p(f"    Still checking ({attempt + 1}/6)... waiting for the page to clear")
            try:
                page.reload(wait_until="domcontentloaded", timeout=timeout)
            except Exception:
                pass
    except Exception as e:
        p(f"    [!] Headed clearance error: {e}")

    if cleared:
        ctx.storage_state(path=STATE_FILE)
        p("  [OK] Session cleared — cookies saved for the rest of this run.")
    else:
        p("  [!] Could not clear the challenge even in a visible window.")

    try:
        ctx.close()
    except Exception:
        pass
    browser.close()
    return cleared


def new_cleared_page(playwright, browser, warm_url=None, timeout=30000, allow_headed=True):
    """Open a fresh context/page and retry until Cloudflare's challenge clears.
    Reuses saved session cookies when available; falls back to a one-time
    visible-browser clearance if quick headless attempts don't get through."""
    state_path = STATE_FILE if os.path.exists(STATE_FILE) else None

    for attempt in range(CHALLENGE_RETRIES):
        ctx_kwargs = dict(
            user_agent=USER_AGENT,
            viewport={"width": 1280, "height": 900},
            locale="en-CA",
        )
        if state_path:
            ctx_kwargs["storage_state"] = state_path
        ctx = browser.new_context(**ctx_kwargs)
        ctx.route("**/*.{mp4,mp3,avi,wmv,woff2,woff}", lambda r: r.abort())
        page = ctx.new_page()
        try:
            page.goto(warm_url or SAVEON_BASE, wait_until="domcontentloaded", timeout=timeout)
            if wait_for_real_content(page, max_wait=25):
                try:
                    ctx.storage_state(path=STATE_FILE)
                except Exception:
                    pass
                return ctx, page
        except Exception:
            pass
        p(f"    [!] Cloudflare challenge attempt {attempt + 1}/{CHALLENGE_RETRIES} failed, retrying...")
        try:
            ctx.close()
        except Exception:
            pass
        state_path = None  # saved cookies (if any) didn't help; try a clean session next
        time.sleep(min(3 * (attempt + 1), 15))

    if allow_headed and playwright is not None:
        p("    [!] Headless retries exhausted — falling back to a one-time visible-browser clearance.")
        if headed_clear_and_save(playwright, timeout=45000):
            return new_cleared_page(playwright, browser, warm_url=warm_url, timeout=timeout, allow_headed=False)

    return None, None


def safe_goto(playwright, browser, page_holder, url, retries=2):
    """Navigate within the existing page; if challenged or errored, get a fresh cleared session."""
    for attempt in range(retries + 1):
        page = page_holder["page"]
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=40000)
            time.sleep(2.5)
            if not is_challenged(page):
                return True
            p("    [!] Hit Cloudflare challenge mid-run, opening fresh session...")
        except Exception as e:
            p(f"    [!] Nav failed: {e}")

        try:
            page_holder["ctx"].close()
        except Exception:
            pass
        ctx, new_page = new_cleared_page(playwright, browser, warm_url=url)
        if new_page is None:
            if attempt == retries:
                return False
            time.sleep(3)
            continue
        page_holder["ctx"] = ctx
        page_holder["page"] = new_page
        return True
    return False


CATEGORY_JS = """
(() => {
    const prefix = window.location.origin;
    const seen = new Set(); const results = [];
    for (const a of document.querySelectorAll('a[href*="/categories/"]')) {
        const slug = a.href
            .replace(prefix, '')
            .replace(/\\/sm\\/[^/]+\\/rsid\\/\\d+/, '')
            .replace('/categories/', '')
            .trim();
        if (!slug || seen.has(slug) || slug.startsWith('category/')) continue;
        if (!slug.includes('-id-')) continue;
        seen.add(slug);
        results.push({ slug, name: a.innerText.trim() });
    }
    return results;
})()
"""


def _leafs_from_cats(all_cats):
    def own_base(slug):
        last = slug.rsplit("/", 1)[-1]
        return re.sub(r"-id-\d+$", "", last)

    parent_bases = set(c["slug"].split("/")[0] for c in all_cats if "/" in c["slug"])
    return [c for c in all_cats if own_base(c["slug"]) not in parent_bases]


def discover_categories(playwright, browser, page_holder, max_attempts=8):
    p("  Discovering categories from /departments (scrolling to load menu)...")

    for attempt in range(max_attempts):
        if not safe_goto(playwright, browser, page_holder, f"{SAVEON_BASE}/departments"):
            continue

        page = page_holder["page"]
        for _ in range(15):
            try:
                page.mouse.wheel(0, 2000)
            except Exception:
                break
            time.sleep(0.4)
        time.sleep(2)

        try:
            all_cats = page.evaluate(CATEGORY_JS)
        except Exception as e:
            p(f"    [!] Category discovery error: {e}")
            all_cats = []

        leafs = _leafs_from_cats(all_cats)
        if leafs:
            p(f"  Found {len(leafs)} leaf categories  ({len(all_cats)} total incl. parents)")
            return [(c["slug"], c["name"]) for c in leafs]

        p(f"    [!] Menu not loaded yet (attempt {attempt + 1}/{max_attempts}), retrying with a fresh session...")
        try:
            page_holder["ctx"].close()
        except Exception:
            pass
        ctx, new_page = new_cleared_page(playwright, browser, warm_url=f"{SAVEON_BASE}/departments")
        if new_page is not None:
            page_holder["ctx"] = ctx
            page_holder["page"] = new_page
        time.sleep(2)

    p("  [!] Could not load the category menu after multiple attempts.")
    return []


def extract_products(page, category=""):
    cards = page.query_selector_all(CARD_SEL)
    results = []
    for card in cards:
        try:
            raw_id = card.get_attribute("data-testid") or ""
            pnum   = raw_id.replace("ProductCardWrapper-", "").lstrip("0") or raw_id

            ps   = card.query_selector_all("p")
            name = ""
            if ps:
                name = re.sub(r",\s*\$[\d.]+.*$", "", ps[0].inner_text()).strip()
            name = name.replace("Open Product Description", "").strip()

            pe      = card.query_selector(PRICE_SEL)
            we      = card.query_selector(WAS_SEL)
            current = pe.inner_text().strip() if pe else ""
            was_raw = we.inner_text().strip() if we else ""
            was     = re.sub(r"^was\s*", "", was_raw, flags=re.IGNORECASE).strip()

            if not name and not pnum:
                continue

            results.append({
                "Category":       category,
                "Product Name":   name,
                "Product Number": pnum,
                "Regular Price":  was if was else current,
                "Promo Price":    current if was else "",
            })
        except Exception:
            continue
    return results


def scrape_all_pages(playwright, browser, page_holder, base_url, category, page_size=30):
    if not safe_goto(playwright, browser, page_holder, base_url):
        return []

    page = page_holder["page"]
    total_items = 0
    try:
        pag_el = page.query_selector('[class*="Pagination"]')
        if pag_el:
            m = re.search(r"of\s+([\d,]+)", pag_el.inner_text())
            if m:
                total_items = int(m.group(1).replace(",", ""))
    except Exception:
        pass

    prods_p1 = extract_products(page, category)
    if not prods_p1 and total_items == 0:
        return []

    if total_items == 0:
        total_items = len(prods_p1)

    total_pages = max(1, -(-total_items // page_size))
    all_prods   = {pr["Product Number"]: pr for pr in prods_p1 if pr["Product Number"]}

    for pg in range(2, total_pages + 1):
        url = f"{base_url}?page={pg}&skip={(pg-1)*page_size}"
        if not safe_goto(playwright, browser, page_holder, url):
            break
        page = page_holder["page"]
        prods = extract_products(page, category)
        if not prods:
            break
        for pr in prods:
            pn = pr["Product Number"]
            if pn and pn not in all_prods:
                all_prods[pn] = pr

    return list(all_prods.values())


def scrape_save_on_foods(playwright):
    p("=" * 60)
    p("  Scraping SAVE-ON-FOODS")
    p("=" * 60)

    browser = playwright.chromium.launch(headless=True)
    all_prods = {}

    ctx, page = new_cleared_page(playwright, browser)
    if page is None:
        p("  [!] Could not get past Cloudflare after multiple attempts.")
        browser.close()
        return []
    page_holder = {"ctx": ctx, "page": page}

    categories = discover_categories(playwright, browser, page_holder)
    if not categories:
        p("  [!] No categories found.")
        try:
            page_holder["ctx"].close()
        except Exception:
            pass
        browser.close()
        return []

    limit_env = os.environ.get("SCRAPE_CATEGORY_LIMIT", "").strip()
    limit = int(limit_env) if limit_env else 0
    if limit:
        categories = categories[:limit]
        p(f"  [TEST MODE] Limiting to {limit} categories.")

    p()
    total_cats = len(categories)

    for i, (slug, cat_name) in enumerate(categories, 1):
        url   = f"{SAVEON_BASE}/categories/{slug}"
        prods = scrape_all_pages(playwright, browser, page_holder, url, cat_name)

        new = 0
        for pr in prods:
            pn = pr["Product Number"]
            if not pn:
                continue
            if pn not in all_prods:
                all_prods[pn] = pr
                new += 1

        p(f"  [{i}/{total_cats}] {cat_name}")
        p(f"    {len(prods)} items  (+{new} new, running total: {len(all_prods):,})")

    p()
    p("  Overlaying promo prices from /promotions...")
    promo_url   = f"{SAVEON_BASE}/promotions"
    promo_total = 0

    if safe_goto(playwright, browser, page_holder, promo_url):
        page = page_holder["page"]
        try:
            pag_el = page.query_selector('[class*="Pagination"]')
            if pag_el:
                m = re.search(r"of\s+([\d,]+)", pag_el.inner_text())
                if m:
                    promo_total = int(m.group(1).replace(",", ""))
        except Exception:
            pass

        promo_pages = max(1, -(-promo_total // 30)) if promo_total else 1
        if limit:
            promo_pages = min(promo_pages, limit)
        p(f"  Promotions: {promo_total:,} items across {promo_pages} pages")

        for pg in range(1, promo_pages + 1):
            url = f"{promo_url}?page={pg}&skip={(pg-1)*30}"
            if pg > 1 and not safe_goto(playwright, browser, page_holder, url):
                break
            page = page_holder["page"]
            prods = extract_products(page, "Sale")
            for pr in prods:
                pn = pr["Product Number"]
                if not pn:
                    continue
                if pn in all_prods:
                    if pr.get("Promo Price"):
                        all_prods[pn]["Promo Price"]  = pr["Promo Price"]
                        all_prods[pn]["Regular Price"] = pr["Regular Price"]
                else:
                    all_prods[pn] = pr

            if pg % 10 == 0 or pg == promo_pages:
                p(f"  Promotions page {pg}/{promo_pages}  (total: {len(all_prods):,})")

    try:
        page_holder["ctx"].close()
    except Exception:
        pass
    browser.close()

    p(f"\n  Save-On-Foods finished — {len(all_prods):,} unique products.\n")
    return list(all_prods.values())


COLS       = ["Category", "Product Name", "Product Number", "Regular Price", "Promo Price"]
COL_WIDTHS = [28, 50, 18, 16, 16]

CAT_COLOURS = [
    "D9E1F2", "E2EFDA", "FCE4D6", "FFF2CC", "EDEDED",
    "D6E4F7", "F2D7EE", "D7F2EE", "F7ECD6", "E8D7F2",
]


def build_excel(products, filepath):
    products = sorted(products, key=lambda x: (x.get("Category", ""),
                                               x.get("Product Name", "")))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Save-On-Foods"

    def bdr(color="BDD7EE"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    last_col = get_column_letter(len(COLS))

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "Save-On-Foods  -  Product Price List"
    c.font      = Font(bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor="D6E4F7")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = (f"Scraped on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}   "
                   f"|   Total products: {len(products):,}")
    c.font      = Font(italic=True, size=9, color="555555")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    for i, col in enumerate(COLS, 1):
        c = ws.cell(row=3, column=i, value=col)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()
    ws.row_dimensions[3].height = 22

    cat_colour_idx = {}
    colour_counter = 0
    prev_cat       = None

    for row_i, pr in enumerate(products, 4):
        cat = pr.get("Category", "")
        if cat not in cat_colour_idx:
            cat_colour_idx[cat] = colour_counter % len(CAT_COLOURS)
            colour_counter += 1

        is_cat_start = (cat != prev_cat)
        prev_cat     = cat
        fill = PatternFill("solid", fgColor=CAT_COLOURS[cat_colour_idx[cat]])

        values = [cat, pr.get("Product Name",""), pr.get("Product Number",""),
                  pr.get("Regular Price",""), pr.get("Promo Price","")]
        for col_i, val in enumerate(values, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border    = bdr("C9C9C9")
            c.fill      = fill
            c.alignment = Alignment(vertical="center")

        ws.cell(row=row_i, column=1).font      = Font(color="444444", size=9, italic=True)
        ws.cell(row=row_i, column=2).alignment = Alignment(horizontal="left",   vertical="center")
        ws.cell(row=row_i, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_i, column=3).font      = Font(name="Courier New", size=9)
        ws.cell(row=row_i, column=4).alignment = Alignment(horizontal="right",  vertical="center")
        ws.cell(row=row_i, column=5).alignment = Alignment(horizontal="right",  vertical="center")
        if pr.get("Promo Price"):
            ws.cell(row=row_i, column=5).font = Font(bold=True, color="C00000")

        if is_cat_start and row_i > 4:
            for col_i in range(1, len(COLS) + 1):
                old = ws.cell(row=row_i, column=col_i).border
                ws.cell(row=row_i, column=col_i).border = Border(
                    left=old.left, right=old.right, bottom=old.bottom,
                    top=Side(style="medium", color="888888"))

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col}{3 + len(products)}"

    wb.save(filepath)
    p(f"  [SAVED] {filepath}")


def main():
    start = time.time()
    p(f"  Output folder : {OUTPUT_DIR}")
    p(f"  Started at    : {datetime.now().strftime('%I:%M %p')}")
    p()

    products = []

    try:
        with sync_playwright() as pw:
            products = scrape_save_on_foods(pw)

        if products:
            p("  Building Excel file...")
            build_excel(products, SAVEON_FILE)
        else:
            p("  [!] No products collected.")

    except KeyboardInterrupt:
        p("\n  Stopped by user.")
    except Exception as e:
        p(f"\n  UNEXPECTED ERROR: {e}")
        import traceback
        traceback.print_exc()
        p()
        p("  Please send the above error to whoever set this up.")

    mins, secs = divmod(int(time.time() - start), 60)
    p()
    p("=" * 60)
    p("  ALL DONE!")
    p("=" * 60)
    if products:
        p(f"  Save-On-Foods    : {len(products):>6,} products")
        p(f"  File saved to    : {SAVEON_FILE}")
    p(f"  Time taken       : {mins}m {secs}s")
    p()

    try:
        input("  Press ENTER to close this window...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
