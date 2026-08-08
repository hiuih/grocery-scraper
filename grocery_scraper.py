#!/usr/bin/env python3
"""
FRESH ST. MARKET PRICE SCRAPER
================================
Scrapes every product from Fresh St. Market (Vancouver House, store 055)
and saves results to Fresh_St_Market_Products.xlsx on your Desktop.

Expected output: ~8,712 unique products across all departments.

HOW TO RUN:
  Double-click "Run Grocery Scraper.bat"  OR  run:
      python grocery_scraper.py
  Takes about 30-40 minutes. Excel file appears on your Desktop when done.
"""

# ── Fix console encoding ──────────────────────────────────────────
import sys, os

try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    elif hasattr(sys.stdout, "buffer"):
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8",
                                      errors="replace", line_buffering=True)
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8",
                                      errors="replace", line_buffering=True)
except Exception:
    pass


def p(msg="", flush=True):
    print(msg, flush=flush)


# ─────────────────────────────────────────────────────────────────
#  Auto-install packages + browser
# ─────────────────────────────────────────────────────────────────
import subprocess

p("=" * 60)
p("  FRESH ST. MARKET PRICE SCRAPER")
p("=" * 60)
p()
p("Checking dependencies...")


def pip_install(pkg):
    r = subprocess.run([sys.executable, "-m", "pip", "install", pkg, "-q"],
                       capture_output=True, text=True)
    if r.returncode != 0:
        subprocess.run([sys.executable, "-m", "pip", "install", pkg])


for pkg in ["playwright", "openpyxl"]:
    try:
        __import__(pkg)
        p(f"  [OK] {pkg}")
    except ImportError:
        p(f"  Installing {pkg}...")
        pip_install(pkg)
        p(f"  [OK] {pkg} installed.")

p("  Checking browser...")
subprocess.run([sys.executable, "-m", "playwright", "install", "chromium", "--with-deps"],
               capture_output=True, text=True)
p("  [OK] Browser ready.")
p()

# ─────────────────────────────────────────────────────────────────
#  Imports
# ─────────────────────────────────────────────────────────────────
try:
    from playwright.sync_api import sync_playwright
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import re, time
    from datetime import datetime
except ImportError as e:
    p(f"  ERROR: {e}")
    try:
        input("  Press ENTER to close...")
    except EOFError:
        pass
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────
#  Desktop path (OneDrive-aware)
# ─────────────────────────────────────────────────────────────────
def find_desktop():
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
        d, _ = winreg.QueryValueEx(key, "Desktop")
        winreg.CloseKey(key)
        if os.path.isdir(d):
            return d
    except Exception:
        pass
    home = os.path.expanduser("~")
    for candidate in [
        os.path.join(home, "Desktop"),
        os.path.join(home, "OneDrive", "Desktop"),
        os.path.join(home, "OneDrive - Personal", "Desktop"),
    ]:
        if os.path.isdir(candidate):
            return candidate
    fb = os.path.join(home, "Desktop")
    os.makedirs(fb, exist_ok=True)
    return fb


OUTPUT_DIR    = find_desktop()
FRESH_ST_FILE = os.path.join(OUTPUT_DIR, "Fresh_St_Market_Products.xlsx")

# ─────────────────────────────────────────────────────────────────
#  Site config
# ─────────────────────────────────────────────────────────────────
FRESH_ST_BASE = "https://www.freshstmarket.com/sm/pickup/rsid/055"

# ─────────────────────────────────────────────────────────────────
#  DOM selectors
# ─────────────────────────────────────────────────────────────────
CARD_SEL  = "[data-testid^='ProductCardWrapper']"
PRICE_SEL = "[class*='ProductPrice--']"
WAS_SEL   = "[class*='ProductWasPrice--']"

# ─────────────────────────────────────────────────────────────────
#  Browser
# ─────────────────────────────────────────────────────────────────
def make_browser(playwright):
    browser = playwright.chromium.launch(headless=True)
    ctx = browser.new_context(
        user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"),
        viewport={"width": 1280, "height": 900},
        locale="en-CA",
    )
    ctx.route("**/*.{mp4,mp3,avi,wmv,woff2,woff}", lambda r: r.abort())
    return browser, ctx


def safe_goto(page, url, retries=2):
    for attempt in range(retries + 1):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=40000)
            time.sleep(2.5)
            return True
        except Exception as e:
            if attempt == retries:
                p(f"    [!] Failed: {e}")
                return False
            time.sleep(3)
    return False


# ─────────────────────────────────────────────────────────────────
#  Category discovery
# ─────────────────────────────────────────────────────────────────
def discover_categories(page):
    p("  Discovering categories from /departments...")
    if not safe_goto(page, f"{FRESH_ST_BASE}/departments"):
        return []

    js = """
    (() => {
        const prefix = 'https://www.freshstmarket.com';
        const seen = new Set(); const results = [];
        for (const a of document.querySelectorAll('a[href*="/categories/"]')) {
            const slug = a.href
                .replace(prefix, '')
                .replace(/\\/sm\\/[^/]+\\/rsid\\/\\d+/, '')
                .replace('/categories/', '')
                .trim();
            if (!slug || seen.has(slug) || slug.startsWith('category/')) continue;
            if (!slug.includes('-id-')) continue;
            seen.add(slug);
            results.push({ slug, name: a.innerText.trim() });
        }
        return results;
    })()
    """
    try:
        all_cats = page.evaluate(js)
    except Exception as e:
        p(f"  [!] Category discovery error: {e}")
        return []

    leafs = [c for c in all_cats if '/' in c['slug']]
    p(f"  Found {len(leafs)} leaf categories  ({len(all_cats)} total incl. parents)")
    return [(c['slug'], c['name']) for c in leafs]


# ─────────────────────────────────────────────────────────────────
#  Product extraction from one loaded page
# ─────────────────────────────────────────────────────────────────
def extract_products(page, category=""):
    cards = page.query_selector_all(CARD_SEL)
    results = []
    for card in cards:
        try:
            raw_id = card.get_attribute("data-testid") or ""
            pnum   = raw_id.replace("ProductCardWrapper-", "").lstrip("0") or raw_id

            ps   = card.query_selector_all("p")
            name = ""
            if len(ps) >= 2:
                name = ps[1].inner_text().strip()
            elif ps:
                name = re.sub(r",\s*\$[\d.]+.*$", "", ps[0].inner_text()).strip()
            name = name.replace("Open Product Description", "").strip()

            pe      = card.query_selector(PRICE_SEL)
            we      = card.query_selector(WAS_SEL)
            current = pe.inner_text().strip() if pe else ""
            was_raw = we.inner_text().strip() if we else ""
            was     = re.sub(r"^was\s*", "", was_raw, flags=re.IGNORECASE).strip()

            if not name and not pnum:
                continue

            results.append({
                "Category":       category,
                "Product Name":   name,
                "Product Number": pnum,
                "Regular Price":  was if was else current,
                "Promo Price":    current if was else "",
            })
        except Exception:
            continue
    return results


# ─────────────────────────────────────────────────────────────────
#  Paginate through all pages of one category URL
# ─────────────────────────────────────────────────────────────────
def scrape_all_pages(page, base_url, category, page_size=30):
    if not safe_goto(page, base_url):
        return []

    total_items = 0
    try:
        pag_el = page.query_selector('[class*="Pagination"]')
        if pag_el:
            m = re.search(r"of\s+([\d,]+)", pag_el.inner_text())
            if m:
                total_items = int(m.group(1).replace(",", ""))
    except Exception:
        pass

    prods_p1    = extract_products(page, category)
    if not prods_p1 and total_items == 0:
        return []

    if total_items == 0:
        total_items = len(prods_p1)

    total_pages = max(1, -(-total_items // page_size))
    all_prods   = {pr["Product Number"]: pr for pr in prods_p1 if pr["Product Number"]}

    for pg in range(2, total_pages + 1):
        url = f"{base_url}?page={pg}&skip={(pg-1)*page_size}"
        if not safe_goto(page, url):
            break
        prods = extract_products(page, category)
        if not prods:
            break
        for pr in prods:
            pn = pr["Product Number"]
            if pn and pn not in all_prods:
                all_prods[pn] = pr

    return list(all_prods.values())


# ─────────────────────────────────────────────────────────────────
#  Main Fresh St. scraper
# ─────────────────────────────────────────────────────────────────
def scrape_fresh_st(playwright):
    p("=" * 60)
    p("  Scraping FRESH ST. MARKET")
    p("=" * 60)

    browser, ctx = make_browser(playwright)
    page = ctx.new_page()
    all_prods = {}

    # Step 1: Discover all leaf categories
    categories = discover_categories(page)
    if not categories:
        p("  [!] No categories found.")
        browser.close()
        return []

    limit_env = os.environ.get("SCRAPE_CATEGORY_LIMIT", "").strip()
    limit = int(limit_env) if limit_env else 0
    if limit:
        categories = categories[:limit]
        p(f"  [TEST MODE] Limiting to {limit} categories.")

    p()
    total_cats = len(categories)

    # Step 2: Scrape each category
    for i, (slug, cat_name) in enumerate(categories, 1):
        url   = f"{FRESH_ST_BASE}/categories/{slug}"
        prods = scrape_all_pages(page, url, cat_name)

        new = 0
        for pr in prods:
            pn = pr["Product Number"]
            if not pn:
                continue
            if pn not in all_prods:
                all_prods[pn] = pr
                new += 1

        p(f"  [{i}/{total_cats}] {cat_name}")
        p(f"    {len(prods)} items  (+{new} new, running total: {len(all_prods):,})")

    # Step 3: Overlay promo prices from /promotions
    p()
    p("  Overlaying promo prices from /promotions...")
    promo_url   = f"{FRESH_ST_BASE}/promotions"
    promo_total = 0

    if safe_goto(page, promo_url):
        try:
            pag_el = page.query_selector('[class*="Pagination"]')
            if pag_el:
                m = re.search(r"of\s+([\d,]+)", pag_el.inner_text())
                if m:
                    promo_total = int(m.group(1).replace(",", ""))
        except Exception:
            pass

        promo_pages = max(1, -(-promo_total // 30)) if promo_total else 1
        if limit:
            promo_pages = min(promo_pages, limit)
        p(f"  Promotions: {promo_total:,} items across {promo_pages} pages")

        for pg in range(1, promo_pages + 1):
            url = f"{promo_url}?page={pg}&skip={(pg-1)*30}"
            if pg > 1 and not safe_goto(page, url):
                break
            prods = extract_products(page, "Sale")
            for pr in prods:
                pn = pr["Product Number"]
                if not pn:
                    continue
                if pn in all_prods:
                    if pr.get("Promo Price"):
                        all_prods[pn]["Promo Price"]  = pr["Promo Price"]
                        all_prods[pn]["Regular Price"] = pr["Regular Price"]
                else:
                    all_prods[pn] = pr

            if pg % 10 == 0 or pg == promo_pages:
                p(f"  Promotions page {pg}/{promo_pages}  (total: {len(all_prods):,})")

    try:
        page.close()
    except Exception:
        pass
    browser.close()

    p(f"\n  Fresh St. finished — {len(all_prods):,} unique products.\n")
    return list(all_prods.values())


# ─────────────────────────────────────────────────────────────────
#  Excel builder
# ─────────────────────────────────────────────────────────────────
COLS       = ["Category", "Product Name", "Product Number", "Regular Price", "Promo Price"]
COL_WIDTHS = [28, 50, 18, 16, 16]

CAT_COLOURS = [
    "D9E1F2", "E2EFDA", "FCE4D6", "FFF2CC", "EDEDED",
    "D6E4F7", "F2D7EE", "D7F2EE", "F7ECD6", "E8D7F2",
]


def build_excel(products, filepath):
    products = sorted(products, key=lambda x: (x.get("Category", ""),
                                               x.get("Product Name", "")))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fresh St. Market"

    def bdr(color="BDD7EE"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    last_col = get_column_letter(len(COLS))

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "Fresh St. Market  -  Product Price List"
    c.font      = Font(bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor="D6E4F7")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = (f"Scraped on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}   "
                   f"|   Total products: {len(products):,}")
    c.font      = Font(italic=True, size=9, color="555555")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    for i, col in enumerate(COLS, 1):
        c = ws.cell(row=3, column=i, value=col)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()
    ws.row_dimensions[3].height = 22

    cat_colour_idx = {}
    colour_counter = 0
    prev_cat       = None

    for row_i, pr in enumerate(products, 4):
        cat = pr.get("Category", "")
        if cat not in cat_colour_idx:
            cat_colour_idx[cat] = colour_counter % len(CAT_COLOURS)
            colour_counter += 1

        is_cat_start = (cat != prev_cat)
        prev_cat     = cat
        fill = PatternFill("solid", fgColor=CAT_COLOURS[cat_colour_idx[cat]])

        values = [cat, pr.get("Product Name",""), pr.get("Product Number",""),
                  pr.get("Regular Price",""), pr.get("Promo Price","")]
        for col_i, val in enumerate(values, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border    = bdr("C9C9C9")
            c.fill      = fill
            c.alignment = Alignment(vertical="center")

        ws.cell(row=row_i, column=1).font      = Font(color="444444", size=9, italic=True)
        ws.cell(row=row_i, column=2).alignment = Alignment(horizontal="left",   vertical="center")
        ws.cell(row=row_i, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_i, column=3).font      = Font(name="Courier New", size=9)
        ws.cell(row=row_i, column=4).alignment = Alignment(horizontal="right",  vertical="center")
        ws.cell(row=row_i, column=5).alignment = Alignment(horizontal="right",  vertical="center")
        if pr.get("Promo Price"):
            ws.cell(row=row_i, column=5).font = Font(bold=True, color="C00000")

        if is_cat_start and row_i > 4:
            for col_i in range(1, len(COLS) + 1):
                old = ws.cell(row=row_i, column=col_i).border
                ws.cell(row=row_i, column=col_i).border = Border(
                    left=old.left, right=old.right, bottom=old.bottom,
                    top=Side(style="medium", color="888888"))

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col}{3 + len(products)}"

    wb.save(filepath)
    p(f"  [SAVED] {filepath}")


# ─────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────
def main():
    start = time.time()
    p(f"  Output folder : {OUTPUT_DIR}")
    p(f"  Started at    : {datetime.now().strftime('%I:%M %p')}")
    p()

    products = []

    try:
        with sync_playwright() as pw:
            products = scrape_fresh_st(pw)

        if products:
            p("  Building Excel file...")
            build_excel(products, FRESH_ST_FILE)
        else:
            p("  [!] No products collected.")

    except KeyboardInterrupt:
        p("\n  Stopped by user.")
    except Exception as e:
        p(f"\n  UNEXPECTED ERROR: {e}")
        import traceback
        traceback.print_exc()
        p()
        p("  Please send the above error to whoever set this up.")

    mins, secs = divmod(int(time.time() - start), 60)
    p()
    p("=" * 60)
    p("  ALL DONE!")
    p("=" * 60)
    if products:
        p(f"  Fresh St. Market : {len(products):>6,} products")
        p(f"  File saved to    : {FRESH_ST_FILE}")
    p(f"  Time taken       : {mins}m {secs}s")
    p()

    try:
        input("  Press ENTER to close this window...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
